#!/usr/bin/env python3
"""
Extracts individual items from the "ideal" LXP backlog xlsx file
and saves each functionality as a separate Markdown file, organized
by module (and optional sub-module).

Usage:
    python3 extract_lxp_backlog.py

Output:
    lxp-backlog/<module-slug>/<functionality-slug>.md
"""

import re
import unicodedata
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "Backlog FEB26 - LXP .xlsx"
OUTPUT_DIR = SCRIPT_DIR / "lxp-backlog"

# Use the first sheet (the one with review columns)
SHEET_NAME = "Backlog por Validar NEES - LXP "

# Column indices (0-based) from the header row
COL_NAME = 0
COL_TYPE = 1
COL_PRIORITY = 2
COL_DESC_EN = 3
COL_DESC_ES = 4
COL_PERTINENTE_GENERAL = 5
COL_OBS_GENERAL = 6
COL_PERTINENTE_TUTORIA = 7
COL_OBS_TUTORIA = 8
COL_PERTINENTE_EVAL = 9
COL_OBS_EVAL = 10
COL_PERTINENTE_APRENDIZAJE = 11
COL_OBS_APRENDIZAJE = 12
COL_PERTINENTE_GESTION = 13
COL_OBS_GESTION = 14


def slugify(text: str) -> str:
    """Convert text to a filesystem-safe slug."""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text[:80].strip("-")


def cell_str(value) -> str:
    """Safely convert a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def build_markdown(item: dict) -> str:
    """Build a Markdown document for a single functionality item."""
    lines = [f"# {item['name']}", ""]

    # Metadata table
    meta = [
        ("Type", item["type"]),
        ("Priority", item["priority"]),
        ("Module", item["module"]),
    ]
    if item.get("sub_module"):
        meta.append(("Sub-Module", item["sub_module"]))

    lines.append("| Field | Value |")
    lines.append("|-------|-------|")
    for field, value in meta:
        if value:
            lines.append(f"| {field} | {value} |")
    lines.append("")

    # Description (English)
    if item["desc_en"]:
        lines.append("## Description (EN)")
        lines.append("")
        lines.append(item["desc_en"])
        lines.append("")

    # Description (Spanish)
    if item["desc_es"]:
        lines.append("## Description (ES)")
        lines.append("")
        lines.append(item["desc_es"])
        lines.append("")

    # Review observations
    review_sections = [
        ("General", item.get("pertinente_general"), item.get("obs_general")),
        ("Tutoria y Formacion", item.get("pertinente_tutoria"), item.get("obs_tutoria")),
        ("Evaluacion", item.get("pertinente_eval"), item.get("obs_eval")),
        ("Aprendizaje", item.get("pertinente_aprendizaje"), item.get("obs_aprendizaje")),
        ("Gestion Escolar", item.get("pertinente_gestion"), item.get("obs_gestion")),
    ]
    has_reviews = any(p or o for _, p, o in review_sections)
    if has_reviews:
        lines.append("## Review Observations")
        lines.append("")
        for section, pertinente, obs in review_sections:
            if pertinente or obs:
                lines.append(f"### {section}")
                if pertinente:
                    lines.append(f"- **Pertinente:** {pertinente}")
                if obs:
                    lines.append(f"- **Observaciones:** {obs}")
                lines.append("")

    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb[SHEET_NAME]

    current_module = ""
    current_sub_module = ""
    total = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = cell_str(row[COL_NAME])
        item_type = cell_str(row[COL_TYPE])

        if not name or not item_type:
            continue

        if item_type == "Module":
            current_module = name
            current_sub_module = ""
            continue

        if item_type == "Sub-Module":
            current_sub_module = name
            continue

        # It's a Functionality — extract and write
        item = {
            "name": name,
            "type": item_type,
            "priority": cell_str(row[COL_PRIORITY]),
            "module": current_module,
            "sub_module": current_sub_module,
            "desc_en": cell_str(row[COL_DESC_EN]),
            "desc_es": cell_str(row[COL_DESC_ES]) if len(row) > COL_DESC_ES else "",
            "pertinente_general": cell_str(row[COL_PERTINENTE_GENERAL]) if len(row) > COL_PERTINENTE_GENERAL else "",
            "obs_general": cell_str(row[COL_OBS_GENERAL]) if len(row) > COL_OBS_GENERAL else "",
            "pertinente_tutoria": cell_str(row[COL_PERTINENTE_TUTORIA]) if len(row) > COL_PERTINENTE_TUTORIA else "",
            "obs_tutoria": cell_str(row[COL_OBS_TUTORIA]) if len(row) > COL_OBS_TUTORIA else "",
            "pertinente_eval": cell_str(row[COL_PERTINENTE_EVAL]) if len(row) > COL_PERTINENTE_EVAL else "",
            "obs_eval": cell_str(row[COL_OBS_EVAL]) if len(row) > COL_OBS_EVAL else "",
            "pertinente_aprendizaje": cell_str(row[COL_PERTINENTE_APRENDIZAJE]) if len(row) > COL_PERTINENTE_APRENDIZAJE else "",
            "obs_aprendizaje": cell_str(row[COL_OBS_APRENDIZAJE]) if len(row) > COL_OBS_APRENDIZAJE else "",
            "pertinente_gestion": cell_str(row[COL_PERTINENTE_GESTION]) if len(row) > COL_PERTINENTE_GESTION else "",
            "obs_gestion": cell_str(row[COL_OBS_GESTION]) if len(row) > COL_OBS_GESTION else "",
        }

        # Build output path: lxp-backlog/<module>/<functionality>.md
        module_slug = slugify(current_module) if current_module else "uncategorized"
        if current_sub_module:
            module_slug = f"{module_slug}/{slugify(current_sub_module)}"

        out_dir = OUTPUT_DIR / module_slug
        out_dir.mkdir(parents=True, exist_ok=True)

        filename = f"{slugify(name)}.md"
        out_path = out_dir / filename
        out_path.write_text(build_markdown(item), encoding="utf-8")
        total += 1

    print(f"Done! Extracted {total} functionalities to {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
